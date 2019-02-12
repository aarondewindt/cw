from typing import Union, Dict, Iterable
from typing.io import BinaryIO
from os import PathLike
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO

from sim_common.object_hierarchies import object_hierarchy_to_tables


def loads(bytes_object) -> Union[Dict, Iterable]:
    """
    """
    pass


def dumps(obj: Union[Dict, Iterable]) -> bytes:
    """
    Returns a bytestring with the object hierarchy serialized into a xlsx/xlsm file.
    """
    # Create byte buffer.
    # Write excel file to the byte buffer.
    # Return the value of the byte buffer.
    buffer = BytesIO()
    dump(obj, buffer)
    return buffer.getvalue()


def load(file: Union[str, BinaryIO, PathLike]) -> Union[Dict, Iterable]:
    """
    """
    pass


def dump(obj: Union[Dict, Iterable], file: Union[str, BinaryIO, PathLike]):
    """
    Dumps an object hierarchy into an xlsx/xlsm file.
    """
    # Create a new workbook and delete the default sheet.
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_font = Font(name='Calibri',
                       size=11,
                       bold=True,
                       italic=False,
                       vertAlign=None,
                       underline='none',
                       strike=False,
                       color='FF000000')

    # Loop through all tables returned by object_hierarchy_to_tables.
    for length, table in object_hierarchy_to_tables(obj).items():
        # If it's a table of scalar elements (aka. length 0), create
        # two columns. The first column has the element name and the
        # second one the value.
        if length == 0:
            worksheet = workbook.create_sheet(f"scalar")
            for i, row in enumerate(table.items()):
                worksheet.append(row)
                worksheet[i+1][0].font = header_font
        else:
            # For any other length create a header with the element names
            # and list the values beneath them.
            worksheet = workbook.create_sheet(f"size {str(length)}")

            # Create the header and set the font
            worksheet.append(tuple(table.keys()))
            for i in range(len(table)):
                worksheet[1][i].font = header_font

            # Loop row by row, and list the values under the headers.
            for row in zip(*[values for _, values in table.items()]):
                worksheet.append(row)

            # Freeze the first row.
            worksheet.freeze_panes = "A2"

        # Adjust the width of the columns to make sure the headers are visible.
        adjust_column_width(worksheet)

    # Create zip archive in with to write the excel data.
    archive = ZipFile(file, 'w', ZIP_DEFLATED, allowZip64=True)

    # Try to write the excel data.
    writer = ExcelWriter(workbook, archive)
    try:
        writer.write_data()
    finally:
        archive.close()


def adjust_column_width(worksheet: Worksheet):
    """
    Adjusts the width of all columns so all of the content is visible.

    :param worksheet: Worksheet object.
    """
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                try:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)*1.4))
                except TypeError:
                    pass
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value
